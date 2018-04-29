# -*- coding: UTF-8 -*-
import openpyxl as xl
from os import path
from config import *
from itertools import combinations
import progressbar


def add_all_permutations(row, new_sheet, counter, all_char_combinations):
    """
    adds all possible replacement combinations for the word given in row. the combinations are being processed,
    and added to the new excel sheet (stored in memory till the sheet is closed).
    """
    current_word = row[0].value.encode("UTF-8")
    for combination in all_char_combinations:
        for char in combination:
            while char in current_word:
                for equivalent_char in HOMOPHONE_HEBREW_CHARS[char]:
                    counter += 1
                    new_sheet["A%d" % counter] = current_word.replace(char, equivalent_char, 1)
                    current_word = current_word.replace(char, equivalent_char, 1)


def create_all_permutations_excel():
    """
    creates a worksheet which will contain all the possible permutations for each word.
    the new excel file will be saved to EXCEL_FILE_PATH with the name NEW_EXCEL_FILE_NAME.
    """
    excel_file = xl.load_workbook(path.join(EXCEL_FILE_PATH, EXCEL_FILE_NAME), read_only=True)
    excel_sheet = excel_file["sheet1"]
    new_excel_file = xl.Workbook()
    new_sheet = new_excel_file.create_sheet()
    counter = 0
    all_chars_combinations = [combination for combination in
                              combinations(HOMOPHONE_HEBREW_CHARS, MAX_WORD_SIZE)]
    rows = excel_sheet['{0}:{1}'.format(excel_sheet.min_row, SEARCH_MAX_ROW)]
    for row in progressbar.progressbar(rows):
        counter += 1
        add_all_permutations(row, new_sheet, counter, all_chars_combinations)
    new_excel_file.save(path.join(EXCEL_FILE_PATH, NEW_EXCEL_FILE_NAME))


if __name__ == '__main__':
    create_all_permutations_excel()
